import os
import threading
from flask import Flask
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from openpyxl import Workbook, load_workbook

# توکن ربات از متغیر محیطی
TOKEN = os.environ.get("TOKEN")
EXCEL_FILE = "contacts.xlsx"
ADMIN_CHAT_IDS = [6441736006, 364551688]

# فایل تصاویر برای هر دپارتمان
DEPARTMENTS = {
    "art": ("هنر و رسانه", "art_media.png", "🎨 دپارتمان هنر و رسانه شامل دوره‌های گرافیک، طراحی و رسانه‌های خلاق است."),
    "computer": ("کامپیوتر", "computer.png", "💻 دپارتمان کامپیوتر شامل دوره‌های برنامه‌نویسی، شبکه، امنیت و نرم‌افزار است."),
    "economy": ("اقتصاد و کوچینگ", "economy_coaching.png", "📈 دپارتمان اقتصاد و کوچینگ شامل آموزش بازارهای مالی، اقتصاد رفتاری و کوچینگ توسعه فردی است."),
    "law": ("حقوق و وکالت", "law_justice.png", "⚖️ دپارتمان حقوق و وکالت شامل آموزش قوانین، وکالت، و حقوق کیفری و مدنی است."),
    "science": ("علمی آزاد", "science_free.png", "🔬 دپارتمان علمی آزاد شامل آموزش‌های عمومی، کاربردی، و دانش آزاد است."),
    "languages": ("زبان‌های خارجی", "languages.png", "🌍 دپارتمان زبان‌های خارجی شامل آموزش زبان انگلیسی، آلمانی، فرانسوی و غیره است.")
}

# اپلیکیشن Flask
app = Flask(__name__)

@app.route('/ping')
def ping():
    return 'pong'

# هندلر /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user

    # اطلاع به ادمین‌ها
    info = (
        f"🆕 کاربر جدید:\n"
        f"👤 {user.first_name} {user.last_name or ''}\n"
        f"🆔 {user.id}\n"
        f"🔗 @{user.username}" if user.username else "بدون یوزرنیم"
    )
    for admin_id in ADMIN_CHAT_IDS:
        await context.bot.send_message(chat_id=admin_id, text=info)

    # ارسال دکمه‌ها
    keyboard = [
        [InlineKeyboardButton(title, callback_data=key)]
        for key, (title, _, _) in DEPARTMENTS.items()
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # دکمه ارسال شماره
    contact_button = ReplyKeyboardMarkup(
        [[KeyboardButton("📱 ارسال شماره من", request_contact=True)]],
        resize_keyboard=True, one_time_keyboard=True
    )

    await update.message.reply_text("سلام! لطفاً دپارتمان مورد نظر را انتخاب کنید:", reply_markup=reply_markup)
    await update.message.reply_text("برای ثبت‌نام، لطفاً شماره تماس خود را ارسال کنید:", reply_markup=contact_button)

# هندلر دکمه‌ها
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    key = query.data
    if key in DEPARTMENTS:
        title, image_file, description = DEPARTMENTS[key]
        caption = f"{description}\n\n☎️ شماره تماس: 03538211100"
        try:
            with open(image_file, "rb") as img:
                await context.bot.send_photo(chat_id=query.message.chat.id, photo=img, caption=caption)
        except FileNotFoundError:
            await query.message.reply_text(f"❌ تصویر دپارتمان «{title}» یافت نشد.")

# هندلر دریافت شماره تماس و ذخیره در فایل اکسل
async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    user = update.message.from_user

    # ساخت فایل اگر وجود ندارد
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["User ID", "First Name", "Last Name", "Username", "Phone Number"])
        wb.save(EXCEL_FILE)

    # ذخیره شماره در فایل
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        user.id,
        user.first_name,
        user.last_name or "",
        f"@{user.username}" if user.username else "",
        contact.phone_number
    ])
    wb.save(EXCEL_FILE)

    await update.message.reply_text("✅ شماره شما با موفقیت ثبت شد.")

    msg = (
        f"📞 شماره جدید:\n"
        f"👤 {user.first_name} {user.last_name or ''}\n"
        f"🆔 {user.id}\n"
        f"🔗 @{user.username}" if user.username else "بدون یوزرنیم\n"
        f"📱 {contact.phone_number}"
    )

    for admin_id in ADMIN_CHAT_IDS:
        await context.bot.send_message(chat_id=admin_id, text=msg)
        try:
            with open(EXCEL_FILE, "rb") as f:
                await context.bot.send_document(chat_id=admin_id, document=f, filename="شماره_دانش‌پذیران.xlsx")
        except Exception as e:
            await context.bot.send_message(chat_id=admin_id, text="❌ خطا در ارسال فایل اکسل.")

# اجرای ربات تلگرام
def run_bot():
    app_telegram = ApplicationBuilder().token(TOKEN).build()
    app_telegram.add_handler(CommandHandler("start", start))
    app_telegram.add_handler(CallbackQueryHandler(button_handler))
    app_telegram.add_handler(MessageHandler(filters.CONTACT, contact_handler))
    print("🤖 ربات در حال اجراست...")
    app_telegram.run_polling()

if __name__ == "__main__":
    # اجرای فلَسک در Thread جدا
    threading.Thread(target=lambda: app.run(host="0.0.0.0", port=8000)).start()
    # اجرای ربات تلگرام
    run_bot()

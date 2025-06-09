from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

EXCEL_FILE = "contacts.xlsx"

# هندلر دریافت شماره تماس و ذخیره در Excel
async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    user = update.message.from_user

    # اگر فایل وجود ندارد، بساز
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["User ID", "First Name", "Last Name", "Username", "Phone Number"])
        wb.save(EXCEL_FILE)

    # افزودن داده جدید
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        user.id,
        user.first_name,
        user.last_name or "",
        f"@{user.username}" if user.username else "",
        contact.phone_number
    ])
    wb.save(EXCEL_FILE)

    # پیام تأیید به کاربر
    await update.message.reply_text("✅ شماره شما با موفقیت ثبت و ذخیره شد.")

    # ارسال مشخصات به ادمین‌ها
    msg = (
        f"📞 شماره جدید:\n"
        f"👤 {user.first_name} {user.last_name or ''}\n"
        f"🆔 {user.id}\n"
        f"🔗 @{user.username}" if user.username else "بدون یوزرنیم\n"
        f"📱 {contact.phone_number}"
    )

    for admin_id in ADMIN_CHAT_IDS:
        await context.bot.send_message(chat_id=admin_id, text=msg)

        # ارسال فایل Excel
        try:
            with open(EXCEL_FILE, "rb") as excel_file:
                await context.bot.send_document(chat_id=admin_id, document=excel_file, filename="شماره_دانش‌پذیران.xlsx")
        except Exception as e:
            await context.bot.send_message(chat_id=admin_id, text="❌ خطا در ارسال فایل اکسل.")
